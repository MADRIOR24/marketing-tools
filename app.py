import streamlit as st
import pandas as pd
import openpyxl
import io

st.set_page_config(page_title="The Marketing Suite", page_icon="📊", layout="wide")

# --- APP NAVIGATION ---
st.sidebar.title("Navigation")
app_mode = st.sidebar.selectbox("Choose a Tool", ["The Survey", "The Segment Creation"])

# --- CONSTANTS ---
BIENNIAL_SHOWS = ["OE", "OI", "SIFER", "FFITA", "FFS", "EFIT", "EUROB", "ICEEE", "INTAI", "EQUDE"]

# List 1: Trigger "Last 5 Editions" Logic
FIVE_YEAR_KEYWORDS = ["LAL"]

# List 2: Trigger "Last 3 Editions" Logic
THREE_YEAR_KEYWORDS = [
    "leads_all_channels", "MSH", "sharers", "Sharer", "Last 3 years exhbiting delegates", 
    "last 2 years registered", "last 3 years registered VIP", "last 3 years registered", 
    "leads", "Last 3 years attended", "Last 3 years attended VIP", "Bounceback",
    "Buyer", "Media", "Last 3 years Reg"
]

# List 3: Trigger "Minus 1 Edition" Logic
MINUS_ONE_KEYWORDS = [
    "Visit to Exhibit Attendee", "Visit to Exhibit non Attendee", "Last year registered VIP", 
    "Last year registered", "Last year attended", "Last year attended VIP", 
    "External Abandon Basket"
]

# Keywords that MUST stay on the current year
CURRENT_YEAR_SHIELDS = ["booked_", "Registration", "Registrations"]

# --- TOOL 1: THE SURVEY ---
if app_mode == "The Survey":
    st.title("📊 The Survey")
    st.write("Extract participant unique codes from Qualtrics URLs.")
    uploaded_file = st.file_uploader("Upload Survey Excel File", type=["xlsx"])
    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        possible_cols = ['Participant_URL', 'URL', 'url', 'Participant URL']
        found_url_col = next((c for c in possible_cols if c in df.columns), None)
        if found_url_col:
            def extract_code(url):
                u = str(url)
                if 'Q_DL=' not in u: return ""
                dl = u.split('Q_DL=')[-1].split('&')[0]
                parts = dl.split('_')
                return "_".join(parts[2:]) if len(parts) > 2 else dl
            df['Participant_Unique_Code'] = df[found_url_col].apply(extract_code)
            cols = list(df.columns)
            cols.remove('Participant_Unique_Code'); idx = cols.index(found_url_col)
            cols.insert(idx + 1, 'Participant_Unique_Code'); df = df[cols]
            st.success("Processing complete!")
            output = io.BytesIO()
            df.to_excel(output, index=False)
            st.download_button("Download Processed File", data=output.getvalue(), file_name="Survey_Processed.xlsx")

# --- TOOL 2: THE SEGMENT CREATION ---
elif app_mode == "The Segment Creation":
    st.title("📁 The Segment Creation")
    st.write("Automate Alpha Code fill-down and Criteria logic.")
    uploaded_file = st.file_uploader("Upload Customer Journey Template", type=["xlsx"])
    
    if uploaded_file:
        wb = openpyxl.load_workbook(uploaded_file)
        sheet = wb.active
        start_row = 4
        m_alpha = str(sheet.cell(row=start_row, column=2).value or "").strip()
        m_year_str = str(sheet.cell(row=start_row, column=3).value or "").strip()
        
        if m_alpha and m_year_str:
            jump = 2 if m_alpha.upper() in BIENNIAL_SHOWS else 1
            st.info(f"Detected {'Biennial' if jump==2 else 'Annual'} logic for {m_alpha}.")
            curr_yr = int(m_year_str)
            
            # MATH PREP
            prev_yr_val = str(curr_yr - jump)
            y3_1, y3_2, y3_3 = str(curr_yr-(jump*3))[-2:], str(curr_yr-(jump*2))[-2:], str(curr_yr-jump)[-2:]
            three_yr_str = f"{y3_1}, {y3_2}, {y3_3}"
            y5_1, y5_2, y5_3, y5_4, y5_5 = str(curr_yr-(jump*5))[-2:], str(curr_yr-(jump*4))[-2:], str(curr_yr-(jump*3))[-2:], str(curr_yr-(jump*2))[-2:], str(curr_yr-jump)[-2:]
            five_yr_str = f"{y5_1}, {y5_2}, {y5_3}, {y5_4}, {y5_5}"

            row = start_row
            while True:
                label_val = sheet.cell(row=row, column=4).value
                if label_val is None: break
                
                label_str = str(label_val).strip()
                label_lower = label_str.lower()
                
                sheet.cell(row=row, column=2).value = m_alpha
                sheet.cell(row=row, column=3).value = m_year_str
                
                # Column G: Segment Name
                clean_lbl = label_str.replace(" ", "_")
                cust_type = sheet.cell(row=row, column=6).value
                sheet.cell(row=row, column=7).value = f"{m_alpha}{m_year_str}_{cust_type}_{clean_lbl}"
                
                # Column I: Criteria Logic
                crit = sheet.cell(row=row, column=9).value
                if crit:
                    lines = str(crit).split('\n')
                    new_lines = []
                    
                    # Logic Priority
                    # 1. Check if the label is Shielded (Current Year)
                    is_shielded = any(s.lower() in label_lower for s in CURRENT_YEAR_SHIELDS)
                    
                    is_5_year = False
                    is_3_year = False
                    is_minus_1 = False

                    # Only run special math if NOT shielded
                    if not is_shielded:
                        is_5_year = any(k.lower() in label_lower for k in FIVE_YEAR_KEYWORDS)
                        if not is_5_year:
                            is_3_year = any(k.lower() in label_lower for k in THREE_YEAR_KEYWORDS)
                        is_minus_1 = any(k.lower() in label_lower for k in MINUS_ONE_KEYWORDS)
                    
                    for line in lines:
                        u_line = line.strip().upper()
                        if u_line.startswith("SHOW ="):
                            new_lines.append(f"SHOW = {m_alpha}")
                        elif u_line.startswith("YEARS ="):
                            if is_5_year:
                                new_lines.append(f"YEARS = {five_yr_str}")
                            elif is_3_year:
                                new_lines.append(f"YEARS = {three_yr_str}")
                            elif is_minus_1:
                                new_lines.append(f"YEARS = {prev_yr_val}")
                            else:
                                # Shielded rows (Registration/booked_) land here
                                new_lines.append(f"YEARS = {m_year_str}")
                        else:
                            new_lines.append(line)
                    sheet.cell(row=row, column=9).value = "\n".join(new_lines)
                row += 1
            
            output = io.BytesIO()
            wb.save(output)
            st.success(f"Processed {row-start_row} rows.")
            st.download_button("Download Generated Segments", data=output.getvalue(), file_name=f"Generated_{m_alpha}_{m_year_str}.xlsx")
